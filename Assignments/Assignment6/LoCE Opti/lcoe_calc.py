#!/usr/bin/env python3
"""
Standalone LCoE Calculator
--------------------------
Reads project parameters from project_list.xlsx, then reads the
optimized layout .txt file to extract turbine (x, y) positions.

Computes:
  • LCoE  (USD/kWh and EUR/MWh)
  • Total Life-Cycle Cost  (TLCC)
  • Net Present Value  (NPV)
  • Profitability Index  (PI)
  • Internal Rate of Return  (IRR)
  • Simple Payback Period (SPP)
  • Cost-breakdown pie chart

Formulas follow TUM "Design of Wind Farms" Lecture 5 (Short et al., 1995).
Cost model mirrors the existing LandBOSSE-equivalent parametric approach.

Usage
-----
  python lcoe_calc.py [layout_txt_path] [xlsx_path]

If no arguments are given, defaults are used (see __main__ block).
The xlsx is always read for project parameters; the txt provides
turbine positions and AEP/cable overrides when available.

xlsx expected columns (any sheet named 'Project' or first sheet):
  - spot_price_eur_mwh   (float)  e.g. 60.0
  - discount_rate        (float)  e.g. 0.036    (fraction, NOT %)
  - lifetime_yr          (int)    e.g. 20
  - turbine_rating_mw    (float)  e.g. 3.3
  - hub_height_m         (float)  e.g. 120.0
  - rotor_diameter_m     (float)  e.g. 156.0
  - turbine_cost_per_kw  (float)  USD/kW
  - foundation_per_turb  (float)  USD
  - roads_per_turb       (float)  USD
  - erection_per_turb    (float)  USD
  - development_fixed    (float)  USD
  - cable_cost_per_km    (float)  USD/km
  - substation_base      (float)  USD
  - substation_per_mw    (float)  USD/MW
  - grid_conn_per_mw     (float)  USD/MW
  - mgmt_fraction        (float)  e.g. 0.604
  - om_fraction          (float)  e.g. 0.012

Any missing column falls back to the hard-coded defaults below.
"""

import re
import sys
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from scipy.optimize import brentq
from pathlib import Path

from landbosse.main_function import run_landbosse # type: ignore

# ── Hard-coded fallback defaults (Denmark / Assignment 6) ─────────────
_DEFAULTS = dict(
    usd_to_eur            = 0.92,
    spot_price_eur_mwh    = 60.0,       # €/MWh  — Denmark avg spot price
    discount_rate         = 0.036,      # real discount rate (fraction)
    lifetime_yr           = 20,
    turbine_rating_mw     = 3.3,        # BAR_BAU_IEA_3.3MW
    hub_height_m          = 120.0,
    rotor_diameter_m      = 156.0,
    turbine_cost_per_kw   = 900.0,      # USD/kW
    foundation_per_turb   = 264_000.0,  # USD
    roads_per_turb        = 33_000.0,   # USD
    erection_per_turb     = 95_000.0,   # USD
    development_fixed     = 314_291.0,  # USD (roughly fixed for site)
    cable_cost_per_km     = 200_000.0,  # USD/km
    substation_base       = 2_000_000.0,
    substation_per_mw     = 22_000.0,   # USD/MW
    grid_conn_per_mw      = 38_400.0,   # USD/MW
    mgmt_fraction         = 0.005,      # fraction of BoS hardware
    om_fraction           = 0.012,      # annual O&M as fraction of CAPEX
)


# ═══════════════════════════════════════════════════════════════════════
# 0.  XLSX READER — load project parameters
# ═══════════════════════════════════════════════════════════════════════
def load_project_params(xlsx_path: str) -> dict:
    """
    Load project parameters from xlsx_path.

    Looks for a sheet named 'Project' (case-insensitive); falls back to
    the first sheet.  Expects a two-column layout: [parameter, value].
    Any missing key falls back to _DEFAULTS.

    Returns a flat dict of parameter values (always numeric where applicable).
    """
    p = Path(xlsx_path)
    if not p.exists():
        print(f"[WARNING] xlsx not found at {xlsx_path!r}. Using all defaults.")
        return dict(_DEFAULTS)

    xl = pd.ExcelFile(xlsx_path)
    sheet_name = next(
        (s for s in xl.sheet_names if "project" in s.lower()),
        xl.sheet_names[0]
    )
    df = xl.parse(sheet_name, header=None)

    # Build a flat dict: first col = key, second col = value
    raw = {}
    for _, row in df.iterrows():
        if len(row) >= 2 and pd.notna(row.iloc[0]) and pd.notna(row.iloc[1]):
            key = str(row.iloc[0]).strip().lower().replace(" ", "_")
            val = row.iloc[1]
            raw[key] = val

    # Merge with defaults; prefer xlsx values
    params = dict(_DEFAULTS)
    key_map = {
        "spot_price_eur_mwh":  "spot_price_eur_mwh",
        "discount_rate":       "discount_rate",
        "lifetime_yr":         "lifetime_yr",
        "turbine_rating_mw":   "turbine_rating_mw",
        "hub_height_m":        "hub_height_m",
        "rotor_diameter_m":    "rotor_diameter_m",
        "turbine_cost_per_kw": "turbine_cost_per_kw",
        "foundation_per_turb": "foundation_per_turb",
        "roads_per_turb":      "roads_per_turb",
        "erection_per_turb":   "erection_per_turb",
        "development_fixed":   "development_fixed",
        "cable_cost_per_km":   "cable_cost_per_km",
        "substation_base":     "substation_base",
        "substation_per_mw":   "substation_per_mw",
        "grid_conn_per_mw":    "grid_conn_per_mw",
        "mgmt_fraction":       "mgmt_fraction",
        "om_fraction":         "om_fraction",
        "usd_to_eur":          "usd_to_eur",
    }
    loaded = []
    for xlsx_key, param_key in key_map.items():
        if xlsx_key in raw:
            val = raw[xlsx_key]
            try:
                params[param_key] = float(val)
                loaded.append(f"  {param_key} = {params[param_key]}")
            except (ValueError, TypeError):
                print(f"[WARNING] Could not parse xlsx value for '{xlsx_key}': {val!r}. Using default.")

    print(f"[INFO] Loaded {len(loaded)} parameter(s) from {xlsx_path!r} (sheet: {sheet_name!r})")
    for l in loaded:
        print(l)

    # Sanity checks — catch common unit mistakes
    if params["discount_rate"] > 1.0:
        print(f"[WARNING] discount_rate={params['discount_rate']} looks like a percentage. "
              f"Dividing by 100 → {params['discount_rate']/100}")
        params["discount_rate"] /= 100.0

    if params["spot_price_eur_mwh"] < 1.0:
        # Probably stored as €/kWh (e.g. 0.06) instead of €/MWh (60)
        print(f"[WARNING] spot_price_eur_mwh={params['spot_price_eur_mwh']} looks like €/kWh. "
              f"Multiplying by 1000 → {params['spot_price_eur_mwh']*1000}")
        params["spot_price_eur_mwh"] *= 1000.0

    if params["om_fraction"] > 0.5:
        print(f"[WARNING] om_fraction={params['om_fraction']} is suspiciously high (should be ~0.012). "
              f"Check units — if it's a percentage, dividing by 100.")
        if params["om_fraction"] > 1.0:
            params["om_fraction"] /= 100.0

    return params


# ═══════════════════════════════════════════════════════════════════════
# 1.  XLSX WRITER — write results back to project_list.xlsx
# ═══════════════════════════════════════════════════════════════════════
def write_results_to_xlsx(xlsx_path: str, results: dict) -> None:
    """
    Append / overwrite a 'Results' sheet in the xlsx with computed metrics.
    """
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    p = Path(xlsx_path)
    if p.exists():
        wb = load_workbook(xlsx_path)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # Remove existing Results sheet
    if "Results" in wb.sheetnames:
        del wb["Results"]
    ws = wb.create_sheet("Results")

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", start_color="2F4F8F")
    unit_font    = Font(italic=True, color="555555")
    good_fill    = PatternFill("solid", start_color="C6EFCE")
    warn_fill    = PatternFill("solid", start_color="FFEB9C")

    rows = [
        ("Metric",                "Value",           "Unit"),
        ("Turbines",              results["n"],       "#"),
        ("Project capacity",      results["project_mw"], "MW"),
        ("AEP",                   round(results["aep_gwh"], 3), "GWh/yr"),
        ("Capacity Factor",       round(results["cap_factor"], 2), "%"),
        ("Cable Run",             round(results["cable_km"], 2), "km"),
        ("",                      "",                ""),
        ("CAPEX",                 round(results["capex_usd"]/1e6, 3), "M USD"),
        ("CAPEX",                 round(results["capex_eur"]/1e6, 3), "M EUR"),
        ("CAPEX per kW",          round(results["capex_per_kw"], 0), "USD/kW"),
        ("Annual OPEX",           round(results["opex_per_yr_eur"]/1e6, 3), "M EUR/yr"),
        ("",                      "",                ""),
        ("FCR",                   round(results["fcr"], 4), "–"),
        ("LCoE",                  round(results["lcoe_usd_kwh"], 4), "USD/kWh"),
        ("LCoE",                  round(results["lcoe_eur_mwh"], 2), "EUR/MWh"),
        ("TLCC",                  round(results["tlcc_eur"]/1e6, 3), "M EUR"),
        ("",                      "",                ""),
        ("Annual Revenue",        round(results["revenue_per_yr_eur"]/1e6, 3), "M EUR/yr"),
        ("NPV",                   round(results["npv_eur"]/1e6, 3), "M EUR"),
        ("PI",                    round(results["pi"], 4), "–"),
        ("SPP",                   round(results["spp_years"], 2), "years"),
        ("IRR",                   round(results["irr"], 2), "%"),
    ]

    for r_idx, (metric, value, unit) in enumerate(rows, start=1):
        ws.cell(r_idx, 1, metric)
        ws.cell(r_idx, 2, value)
        ws.cell(r_idx, 3, unit)

    # Header styling
    for col in range(1, 4):
        c = ws.cell(1, col)
        c.font  = header_font
        c.fill  = header_fill
        c.alignment = Alignment(horizontal="center")

    # Highlight PI row
    pi_row = next(i+1 for i, (m, _, _) in enumerate(rows) if m == "PI")
    pi_val = results["pi"]
    fill = good_fill if pi_val > 1.0 else warn_fill
    for col in range(1, 4):
        ws.cell(pi_row, col).fill = fill

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12

    wb.save(xlsx_path)
    print(f"\n  Results written → {xlsx_path!r}  (sheet: 'Results')")


# ═══════════════════════════════════════════════════════════════════════
# 2.  PARSER — read layout txt → list of (x, y)
# ═══════════════════════════════════════════════════════════════════════
def parse_layout(filepath: str) -> tuple[list[tuple[float, float]], dict]:
    """
    Parse an optimized-layout text file.

    Returns
    -------
    positions : list of (x_m, y_m)  turbine coordinates
    meta      : dict with keys 'aep_gwh', 'n_turbines',
                'substation_xy', 'cable_km' (when available)
    """
    positions = []
    meta = {}

    with open(filepath, "r") as f:
        for line in f:
            line = line.strip()

            m = re.match(r"AEP\s*:\s*([\d.]+)\s*GWh", line)
            if m:
                meta["aep_gwh"] = float(m.group(1))

            m = re.match(r"Substation.*x\s*=\s*([\d.]+)\s*m.*y\s*=\s*([\d.]+)\s*m", line)
            if m:
                meta["substation_xy"] = (float(m.group(1)), float(m.group(2)))

            m = re.match(r"Total cable run.*?([\d.]+)\s*km", line)
            if m:
                meta["cable_km"] = float(m.group(1))

            m = re.match(r"\s*Turbine\s+\d+:\s*x\s*=\s*([\d.]+)\s*m.*y\s*=\s*([\d.]+)\s*m", line)
            if m:
                positions.append((float(m.group(1)), float(m.group(2))))

    meta["n_turbines"] = len(positions)
    return positions, meta


# ═══════════════════════════════════════════════════════════════════════
# 3.  COST + FINANCIAL MODEL
# ═══════════════════════════════════════════════════════════════════════
def compute_lcoe(
    positions: list[tuple[float, float]],
    aep_gwh: float,
    params: dict,
    cable_km: float | None = None,
    substation_xy: tuple[float, float] | None = None,
    show_plot: bool = True,
    save_plot: str | None = None,
    xlsx_path: str | None = None,
) -> dict:
    """
    Compute LCoE and related financial indicators.

    Parameters
    ----------
    positions      : list of (x, y) turbine coordinates in metres
    aep_gwh        : Annual Energy Production in GWh/yr
    params         : project parameter dict (from load_project_params)
    cable_km       : total cable run in km (estimated if None)
    substation_xy  : (x, y) of substation in metres (estimated if None)
    show_plot      : whether to plt.show() the pie chart
    save_plot      : filepath to save the pie chart (None = don't save)
    xlsx_path      : if given, write Results sheet back to this xlsx

    Returns
    -------
    results : dict with all computed metrics
    """
    # Unpack params
    USD_TO_EUR         = params["usd_to_eur"]
    SPOT_PRICE_EUR_MWH = params["spot_price_eur_mwh"]
    DISCOUNT_RATE      = params["discount_rate"]
    LIFETIME_YR        = int(params["lifetime_yr"])
    TURBINE_RATING_MW  = params["turbine_rating_mw"]

    TURBINE_COST_PER_KW  = params["turbine_cost_per_kw"]
    FOUNDATION_PER_TURB  = params["foundation_per_turb"]
    ROADS_PER_TURB       = params["roads_per_turb"]
    ERECTION_PER_TURB    = params["erection_per_turb"]
    DEVELOPMENT_FIXED    = params["development_fixed"]
    CABLE_COST_PER_KM    = params["cable_cost_per_km"]
    SUBSTATION_BASE      = params["substation_base"]
    SUBSTATION_PER_MW    = params["substation_per_mw"]
    GRID_CONN_PER_MW     = params["grid_conn_per_mw"]
    MGMT_FRACTION        = params["mgmt_fraction"]
    OandM_FRACTION       = params["om_fraction"]

    n  = len(positions)
    xs = np.array([p[0] for p in positions])
    ys = np.array([p[1] for p in positions])

    project_MW  = n * TURBINE_RATING_MW
    project_kW  = project_MW * 1e3
    aep_kwh     = aep_gwh * 1e6   # GWh → kWh
    aep_mwh     = aep_gwh * 1e3   # GWh → MWh

    cap_factor  = aep_gwh * 1e3 / (project_MW * 8760) * 100   # %

    # ── Estimate cable length if not provided ──────────────────────
    if cable_km is None:
        if substation_xy is not None:
            sx, sy = substation_xy
        else:
            sx, sy = xs.mean(), ys.min()
        dists   = np.sqrt((xs - sx)**2 + (ys - sy)**2)
        cable_km = dists.sum() / 1000.0 * 1.15

    # ── Individual cost items (USD) ────────────────────────────────
    turbine_cost = project_kW  * TURBINE_COST_PER_KW
    foundation   = n           * FOUNDATION_PER_TURB
    roads        = n           * ROADS_PER_TURB
    erection     = n           * ERECTION_PER_TURB
    development  = DEVELOPMENT_FIXED
    cable_cost   = cable_km    * CABLE_COST_PER_KM
    substation   = SUBSTATION_BASE + project_MW * SUBSTATION_PER_MW
    grid_conn    = project_MW  * GRID_CONN_PER_MW

    bos_hardware = (foundation + roads + erection + development
                    + cable_cost + substation + grid_conn)
    management   = MGMT_FRACTION * bos_hardware
    bos_total    = bos_hardware + management
    capex        = turbine_cost + bos_total
    capex_per_kw = capex / project_kW
    capex_eur    = capex * USD_TO_EUR

    opex_per_yr     = OandM_FRACTION * capex
    opex_per_yr_eur = opex_per_yr * USD_TO_EUR

    # ── Financial metrics ──────────────────────────────────────────
    d = DISCOUNT_RATE
    N = LIFETIME_YR

    fcr     = d * (1 + d)**N / ((1 + d)**N - 1)
    annuity = 1.0 / fcr

    # LCoE = (CAPEX · FCR + OPEX) / AEP
    lcoe_usd_kwh = (capex * fcr + opex_per_yr) / aep_kwh
    lcoe_eur_mwh = lcoe_usd_kwh * 1e3 * USD_TO_EUR

    # TLCC = CAPEX + OPEX · annuity
    tlcc_eur = capex_eur + opex_per_yr_eur * annuity

    # Revenue and NPV
    revenue_per_yr_eur = aep_mwh * SPOT_PRICE_EUR_MWH
    net_annual_eur     = revenue_per_yr_eur - opex_per_yr_eur
    npv_eur            = net_annual_eur * annuity - capex_eur

    # PI = 1 + NPV / I
    pi = 1.0 + npv_eur / capex_eur

    # Simple Payback Period
    spp_years = capex_eur / net_annual_eur if net_annual_eur > 0 else float("inf")

    # IRR
    def npv_func(r):
        if abs(r) < 1e-12:
            return net_annual_eur * N - capex_eur
        ann = ((1 + r)**N - 1) / (r * (1 + r)**N)
        return net_annual_eur * ann - capex_eur

    try:
        irr = brentq(npv_func, -0.5, 10.0) * 100
    except ValueError:
        irr = float("nan")

    # ── Print summary ──────────────────────────────────────────────
    print("=" * 62)
    print("  LCoE & Financial Summary")
    print("=" * 62)
    print(f"  Turbines          : {n}")
    print(f"  Rating            : {TURBINE_RATING_MW:.1f} MW  →  {project_MW:.1f} MW total")
    print(f"  AEP               : {aep_gwh:.3f} GWh/yr")
    print(f"  Capacity Factor   : {cap_factor:.2f} %")
    print(f"  Cable Run         : {cable_km:.2f} km")
    print(f"  Spot Price        : {SPOT_PRICE_EUR_MWH:.1f} EUR/MWh")
    print(f"  Discount Rate     : {d*100:.2f} %")
    print(f"  Lifetime          : {N} yr")
    print("-" * 62)
    print(f"  Turbine Cost      : ${turbine_cost:>14,.0f}")
    print(f"  Foundation        : ${foundation:>14,.0f}")
    print(f"  Roads             : ${roads:>14,.0f}")
    print(f"  Erection          : ${erection:>14,.0f}")
    print(f"  Development       : ${development:>14,.0f}")
    print(f"  Collection Cable  : ${cable_cost:>14,.0f}")
    print(f"  Substation        : ${substation:>14,.0f}")
    print(f"  Grid Connection   : ${grid_conn:>14,.0f}")
    print(f"  Management        : ${management:>14,.0f}")
    print(f"  BoS Total         : ${bos_total:>14,.0f}")
    print(f"  ─────────────────────────────────────────")
    print(f"  CAPEX Total       : ${capex:>14,.0f}   ({capex_per_kw:.0f} $/kW)")
    print(f"  OPEX / yr         : ${opex_per_yr:>14,.0f}")
    print("-" * 62)
    print(f"  FCR               : {fcr:.4f}")
    print(f"  LCoE              : {lcoe_usd_kwh:.4f} USD/kWh")
    print(f"  LCoE              : {lcoe_eur_mwh:.2f} EUR/MWh")
    print(f"  TLCC              : €{tlcc_eur:>14,.0f}")
    print("-" * 62)
    print(f"  Annual Revenue    : €{revenue_per_yr_eur:>14,.0f}")
    print(f"  NPV               : €{npv_eur:>14,.0f}")
    print(f"  PI                : {pi:.4f}")
    print(f"  SPP               : {spp_years:.2f} years")
    print(f"  IRR               : {irr:.2f} %")
    print("=" * 62)

    results = dict(
        n=n, project_mw=project_MW, aep_gwh=aep_gwh, cap_factor=cap_factor,
        cable_km=cable_km, capex_usd=capex, capex_eur=capex_eur,
        capex_per_kw=capex_per_kw, opex_per_yr_eur=opex_per_yr_eur,
        fcr=fcr, lcoe_usd_kwh=lcoe_usd_kwh, lcoe_eur_mwh=lcoe_eur_mwh,
        tlcc_eur=tlcc_eur, revenue_per_yr_eur=revenue_per_yr_eur,
        npv_eur=npv_eur, pi=pi, spp_years=spp_years, irr=irr,
    )

    # ── Write results back to xlsx ─────────────────────────────────
    if xlsx_path:
        write_results_to_xlsx(xlsx_path, results)

    # ── Pie chart ──────────────────────────────────────────────────
    labels = [
        "Turbines", "Foundation", "Roads", "Erection",
        "Development", "Collection Cable", "Substation",
        "Grid Connection", "Management",
    ]
    values = [
        turbine_cost, foundation, roads, erection,
        development, cable_cost, substation, grid_conn, management,
    ]
    colors = [
        "#2196F3", "#FF9800", "#795548", "#4CAF50",
        "#9C27B0", "#F44336", "#00BCD4", "#FFEB3B", "#607D8B",
    ]

    fig, ax = plt.subplots(figsize=(10, 8))
    wedges, texts, autotexts = ax.pie(
        values,
        labels=labels,
        autopct=lambda pct: f"{pct:.1f}%\n(${pct / 100 * capex / 1e6:.1f}M)",
        colors=colors,
        startangle=140,
        pctdistance=0.75,
        wedgeprops=dict(edgecolor="white", linewidth=1.5),
    )
    for t in autotexts:
        t.set_fontsize(8)
    for t in texts:
        t.set_fontsize(9)

    ax.set_title(
        f"CAPEX Breakdown — {n} × {TURBINE_RATING_MW:.1f} MW  "
        f"| LCoE = {lcoe_eur_mwh:.1f} €/MWh  "
        f"| PI = {pi:.3f}",
        fontsize=13, fontweight="bold", pad=20,
    )

    plt.tight_layout()
    if save_plot:
        plt.savefig(save_plot, dpi=200, bbox_inches="tight")
        print(f"\n  Pie chart saved → {save_plot}")
    if show_plot:
        plt.show()
    else:
        plt.close(fig)

    return results


# ═══════════════════════════════════════════════════════════════════════
# 4.  CLI entry-point
# ═══════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    # ── Paths ──────────────────────────────────────────────────────
    DEFAULT_LAYOUT = (
        "Assignments/Assignment6/LoCE Opti/optimizedLayout/"
        "BAR_BAU_IEA_3.3MW_16_layout.txt"
    )
    DEFAULT_XLSX = (
        "/home/lavender/Studies/Design of Wind Farms/Assignments/"
        "Assignment6/LoCE Opti/project_list.xlsx"
    )

    layout_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_LAYOUT
    xlsx_path   = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_XLSX

    # ── Load project parameters from xlsx ─────────────────────────
    params = load_project_params(xlsx_path)

    # ── Parse layout txt ──────────────────────────────────────────
    positions, meta = parse_layout(layout_path)

    if not positions:
        print(f"ERROR: no turbine positions found in {layout_path}")
        sys.exit(1)

    print(f"\nParsed {meta['n_turbines']} turbines from {layout_path!r}")

    # ── Run financial model ────────────────────────────────────────
    results = compute_lcoe(
        positions,
        aep_gwh       = meta.get("aep_gwh", 0.0),
        params        = params,
        cable_km      = meta.get("cable_km"),
        substation_xy = meta.get("substation_xy"),
        show_plot     = True,
        save_plot     = "lcoe_pie_chart.png",
        xlsx_path     = xlsx_path,   # write Results sheet back
    )